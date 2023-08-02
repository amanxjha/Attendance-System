import pandas as pd #panda module
import os #os module
from datetime import datetime
import openpyxl as op # for output files

# for mailing purpose
import smtplib 
from os.path import basename
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.base import MIMEBase
from email import encoders

os.system("cls") # clear the screen

os.chdir(r'C:\Users\Aman Jha\Documents\GitHub\Attendance-System') # changing directory to present directory

#sending mail function
def sendMail(fromaddr, frompasswd, toaddr, msg_subject, msgBody, filePath): 
    try:
        msg = MIMEMultipart()
        print("[+] Message Object Created")
    except:
        print("[-] Error in Creating Message Object")
        return

    msg['From'] = fromaddr

    msg['To'] = toaddr

    msg['Subject'] = msg_subject

    body = msgBody

    msg.attach(MIMEText(body, 'plain'))

    fileName = filePath
    attachment = open(fileName, "rb")

    p = MIMEBase('application', 'octet-stream')

    p.set_payload((attachment).read())

    encoders.encode_base64(p)

    p.add_header('Content-Disposition', "attachment; filename= %s" % fileName)

    try:
        msg.attach(p)
        print("[+] File Attached")
    except:
        print("[-] Error in Attaching file")
        return

    try:
        #s = smtplib.SMTP('smtp.gmail.com', 587)
        s = smtplib.SMTP('stud.iitp.ac.in', 587)
        print("[+] SMTP Session Created")
    except:
        print("[-] Error in creating SMTP session")
        return

    s.starttls()

    try:
        s.login(fromaddr, frompasswd)
        print("[+] Login Successful")
    except:
        print("[-] Login Failed")

    text = msg.as_string()

    try:
        s.sendmail(fromaddr, toaddr, text)
        print("[+] Mail Sent successfully")
    except:
        print('[-] Mail not sent')

    s.quit()

def attendance_report():
    try:
        inputf=pd.read_csv(r'input_registered_students.csv') # opening the input file in read mode with help of pandas
        inputf.dropna() # dropping null rows
    except:
        print("Can't open the input_registered_students.csv file.")
        exit()

    sz=len(inputf) # number of registered students
    tdict={} # dictionary of roll no and their name

    for i in range(sz):
        tdict[inputf.loc[i,"Roll No"]]=inputf.loc[i,"Name"]
    
    try:
        inpf2=pd.read_csv(r'input_attendance.csv') # opening the input file in read mode with help of pandas
        inpf2=inpf2.dropna() # dropping null rows
    except:
        print("Can't open the input_attendance.csv file.")
        exit()
    
    inpf2["Timestamp"]=pd.to_datetime(inpf2["Timestamp"],dayfirst=1) # converting string to datetime
    
    try:
        # getting columns as list
        timestamp=list(inpf2["Timestamp"])
        attendance=list(inpf2["Attendance"])
        roll_no=list(inputf["Roll No"])
    except:
        print("Cant convert to list at line 111")
        exit()

    # setting the time interval in which attendance is valid
    inter=["14:00","15:00"]
    inter[0]=datetime.strptime(inter[0],"%H:%M")
    inter[1]=datetime.strptime(inter[1],"%H:%M")

    # making a list of roll numbers from attendance file
    att_roll=[]
    try:
        for i in attendance:
            li=i.split(" ")
            att_roll.append(li[0])
    except:
        print("Unable to store just the marked roll numbers at line 126.")

    actual_dates=[] # Dates on which lectures were taken
    tot_att_cnt={} # it is a dictionary of type {string:list containing two elements of which first elements gives us total attendance count and other gives us invalid count}

    prev=0
    for i in range(len(timestamp)):
        if timestamp[i].isoweekday()==1 or timestamp[i].isoweekday()==4: # only mondays amd thursdays

            if prev!=timestamp[i].date(): # only unique dates list
                try:
                    # initialization of total attendance count dictionary
                    tot_att_cnt[timestamp[i].date()]={}
                    for j in range(sz):
                        tot_att_cnt[timestamp[i].date()][roll_no[j]]=[0,0]
                except:
                    print("Unable to initialize total attendance count")

                actual_dates.append(timestamp[i].date()) # the dates on which lecture were taken

            prev=timestamp[i].date()

    tot_lec=len(actual_dates) # number of total lectures taken

    for i in range(sz): # for all the registered students in master list
        tot=0; inv=0; prev=0 # tot is to keep the total count, inv is to keep the invalid count, prev is to keep the previous date

        for j in range(len(att_roll)): # for all the attendance marked

            if roll_no[i]==att_roll[j]: # if the roll number of the student matches with the roll no in attendance list

                if timestamp[j].isoweekday()==1 or timestamp[j].isoweekday()==4: # only mondays amd thursdays

                    if prev!=timestamp[j].date(): # if date changes then fill in the details in tot_att_cnt and reinitialize tot, inv, prev
                        if prev!=0:
                            try:
                                tot_att_cnt[prev][roll_no[i]][0]=tot
                                tot_att_cnt[prev][roll_no[i]][1]=inv
                            except:
                                print("failed to store values at line 165.")
                        tot=0; inv=0
                        prev=timestamp[j].date()

                    tot+=1 # increasing the total count in any case

                    if timestamp[j].time()<inter[0].time() or timestamp[j].time()>inter[1].time(): # ioncreasing the invalid count in case it is marked in invalid time period
                        inv+=1
        
        if prev!=0: # putting values for last case
            tot_att_cnt[prev][roll_no[i]][0]=tot
            tot_att_cnt[prev][roll_no[i]][1]=inv
        
    os.chdir(r'C:\Users\Aman Jha\Documents\GitHub\Attendance-System\output') # changing directory to output file for output
    main={} # making a dictionary of {string: list of all  the dates and whether they are present or absent} # this is for the attendance_report_consolidated part

    for i in range(sz): # for all the registered students in master list
        main[roll_no[i]]=[] # initializing with an empty list (this is for the last file)
        str=roll_no[i]+".xlsx" # name of the file for individual list

        tempfi=op.Workbook() # making a workbook
        opsheet=tempfi.active # creating a sheet
        opsheet['A1'].value="Date" # putting headers
        opsheet['B1'].value="Roll"
        opsheet['C1'].value="Name"
        opsheet['D1'].value="Attendance"
        opsheet['E1'].value="Real"
        opsheet['F1'].value="Duplicate"
        opsheet['G1'].value="Invalid"
        opsheet['H1'].value="Absent"
        opsheet['B2'].value=roll_no[i]
        opsheet['C2'].value=tdict[roll_no[i]]

        for j in range(len(actual_dates)):
            tot=tot_att_cnt[actual_dates[j]][roll_no[i]][0] # total
            inv=tot_att_cnt[actual_dates[j]][roll_no[i]][1] # invalid

            opsheet.cell(row=j+3,column=1).value=actual_dates[j]
            opsheet.cell(row=j+3,column=4).value=tot

            if tot-inv>0: # case when student is present on that day
                opsheet.cell(row=j+3,column=5).value=1 # real can be either 1(present) or 0(absent)
                opsheet.cell(row=j+3,column=6).value=tot-inv-1 # count of duplicate
                opsheet.cell(row=j+3,column=8).value=0 # absent
                main[roll_no[i]].append('P')
            else: # case when student is absent on that day
                opsheet.cell(row=j+3,column=5).value=0 # real can be either 1(present) or 0(absent)
                opsheet.cell(row=j+3,column=6).value=0 # count of duplicate is also 0
                opsheet.cell(row=j+3,column=8).value=1 # absent
                main[roll_no[i]].append('A')

            opsheet.cell(row=j+3,column=7).value=inv # invalid
        tempfi.save(str)

    # last file
    mainfi=op.Workbook()
    opsheetm=mainfi.active
    opsheetm['A1'].value="Roll"
    opsheetm['B1'].value="Name"

    for i in range(len(actual_dates)):
        opsheetm.cell(row=1,column=i+3).value=actual_dates[i] # creating dates header
    
    opsheetm.cell(row=1,column=3+len(actual_dates)).value="Actual Lecture Taken"
    opsheetm.cell(row=1,column=4+len(actual_dates)).value="Total Real"
    opsheetm.cell(row=1,column=5+len(actual_dates)).value="% Attendance"

    for i in range(sz):
        opsheetm.cell(row=i+2,column=1).value=roll_no[i]
        opsheetm.cell(row=i+2,column=2).value=tdict[roll_no[i]]

        col=3; cnt=0 # column no, and count of present day
        for j in main[roll_no[i]]:
            opsheetm.cell(row=i+2,column=col).value=j
            if(j=='P'):
                cnt+=1
            col+=1
        opsheetm.cell(row=i+2,column=col).value=tot_lec
        opsheetm.cell(row=i+2,column=col+1).value=cnt
        opsheetm.cell(row=i+2,column=col+2).value=round((cnt/tot_lec)*100,2) # rounding off to 2 digit
    
    mainfi.save("attendance_report_consolidated.xlsx")

    # mailing attendance report consolidated
    FROM_ADDR = "aman_2001ee22@iitp.ac.in" # my mail id
    FROM_PASSWD = "changeme" # my password
    to_addr="changemeaswell" # to the person whom i am sending the mail

    Subject="Attendance Report" # sybject of the mail
    content="PFA attendance_report_consolidated.\nBest regards." # body of the file

    file_path="attendance_report_consolidated.xlsx" # the file which we have to attach

    sendMail(FROM_ADDR, FROM_PASSWD, to_addr, Subject, content, file_path) # send the mail
    
attendance_report() # call the function